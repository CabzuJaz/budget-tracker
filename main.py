import os
from datetime import datetime, timedelta
from openpyxl import Workbook

BASE_DIR = os.path.dirname(os.path.abspath(__file__))

FILES = {
    "Cash": "cash.txt",
    "Ralph_CC": "ralph_cc.txt",
    "Jazz_CC": "jazz_cc.txt",
    "Spaylater": "spaylater.txt"
}

# ======================
# UTIL
# ======================

def parse_date(date_str):
    return datetime.strptime(date_str, "%b %d, %Y")

def format_date(date_obj):
    return date_obj.strftime("%b %d, %Y")

def ensure_file(filename):
    path = os.path.join(BASE_DIR, filename)
    if not os.path.exists(path):
        open(path, "w").close()
    return path

def prev_month(y, m):
    return (y-1, 12) if m == 1 else (y, m-1)

def next_month(y, m):
    return (y+1, 1) if m == 12 else (y, m+1)

# ======================
# CYCLE LOGIC
# ======================

def get_cycle_range(category, date):
    y, m, d = date.year, date.month, date.day

    if category in ["Ralph_CC", "Spaylater"]:
        if d <= 25:
            py, pm = prev_month(y, m)
            return datetime(py, pm, 26), datetime(y, m, 25)
        else:
            ny, nm = next_month(y, m)
            return datetime(y, m, 26), datetime(ny, nm, 25)

    if category == "Jazz_CC":
        if d <= 15:
            py, pm = prev_month(y, m)
            return datetime(py, pm, 16), datetime(y, m, 15)
        else:
            ny, nm = next_month(y, m)
            return datetime(y, m, 16), datetime(ny, nm, 15)

def get_due_date(category, cycle_end):
    if category in ["Ralph_CC", "Spaylater"]:
        return cycle_end + timedelta(days=20)
    if category == "Jazz_CC":
        return datetime(cycle_end.year + (cycle_end.month // 12),
                        (cycle_end.month % 12) + 1, 5)

# ======================
# ADD EXPENSE
# ======================

def add_expense():
    try:
        date_str = input("Date (Apr 30, 2026): ")
        txn_date = parse_date(date_str)

        category = input("Category (Cash, Ralph_CC, Jazz_CC, Spaylater): ")
        if category not in FILES:
            print("Invalid category.")
            return

        desc = input("Description: ")
        amount = float(input("Amount: "))

        file_path = ensure_file(FILES[category])

        if category == "Cash":
            with open(file_path, "a") as f:
                f.write(f"{date_str}|{desc}|-{amount}\n")
        else:
            start, end = get_cycle_range(category, txn_date)
            due = get_due_date(category, end)

            with open(file_path, "a") as f:
                f.write(f"{date_str}|{desc}|{amount}|{format_date(due)}|UNPAID\n")

        print("✅ Saved")

    except Exception as e:
        print("Error:", e)

# ======================
# MARK PAID
# ======================

def mark_paid():
    category = input("Category: ")
    file_path = ensure_file(FILES[category])

    lines = open(file_path).readlines()

    for i, line in enumerate(lines):
        print(f"{i}: {line.strip()}")

    idx = int(input("Select index to mark PAID: "))
    parts = lines[idx].strip().split("|")

    if len(parts) >= 5:
        parts[4] = "PAID"
        lines[idx] = "|".join(parts) + "\n"

    with open(file_path, "w") as f:
        f.writelines(lines)

    print("✅ Marked as PAID")

# ======================
# CURRENT CYCLE TOTAL
# ======================

def get_current_cycle_total(category):
    today = datetime.today()
    start, end = get_cycle_range(category, today)

    total = 0

    with open(ensure_file(FILES[category])) as f:
        for line in f:
            parts = line.strip().split("|")
            txn_date = parse_date(parts[0])
            amount = float(parts[2])
            status = parts[4] if len(parts) > 4 else "UNPAID"

            if start <= txn_date <= end and status == "UNPAID":
                total += amount

    return total, start, end

# ======================
# DASHBOARD
# ======================

def show_dashboard():
    print("\n===== UPCOMING BILLS =====")

    today = datetime.today()
    data = []

    for cat in ["Ralph_CC", "Jazz_CC", "Spaylater"]:
        total, start, end = get_current_cycle_total(cat)
        due = get_due_date(cat, end)
        days = (due - today).days

        data.append((cat, total, due, days))

    data.sort(key=lambda x: x[2])

    for d in data:
        print(f"\n{d[0]}")
        print(f"Due: {format_date(d[2])}")
        print(f"Amount: {d[1]}")
        print(f"Days Left: {d[3]}")

    print("\n MOST URGENT:", data[0][0])

# ======================
# EXPORT EXCEL
# ======================

def export_excel():
    wb = Workbook()

    for category in FILES:
        ws = wb.create_sheet(title=category)

        ws.append(["Date", "Description", "Amount", "Due", "Status"])

        with open(ensure_file(FILES[category])) as f:
            for line in f:
                ws.append(line.strip().split("|"))

    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    path = os.path.join(BASE_DIR, "budget_report.xlsx")
    wb.save(path)

    print("✅ Exported:", path)

# ======================
# MENU
# ======================

def menu():
    while True:
        print("\n==== BUDGET TRACKER ====")
        print("1. Add Expense")
        print("2. Mark Paid")
        print("3. Dashboard")
        print("4. Export Excel")
        print("5. Exit")

        c = input("Choose: ")

        if c == "1":
            add_expense()
        elif c == "2":
            mark_paid()
        elif c == "3":
            show_dashboard()
        elif c == "4":
            export_excel()
        elif c == "5":
            break

if __name__ == "__main__":
    menu()